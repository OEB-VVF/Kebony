"""
Kebony KPI Governance Template — v2.0

Rebuilt per feedback (2026-04-21):
- One sheet per domain (Finance / Sales / Supply Chain / Manufacturing /
  Quality / Inventory / Safety / ESG / Marketing / Product Dev / HR /
  Certification)
- Only derived KPIs (ratios / computed). No elementary data rows.
- Axes of Analysis = dedicated reference sheet + per-KPI checkbox columns
  ("which dimensions of the cube does this KPI slice on?")
- Cover sheet with signoff block per domain head
- Legend & conventions

Axes (from Analytical Accounting Architecture + Reporting - Cube Architecture):
  Entity · Time · Phase · Product · Geography (Country) · Sales Area (Global) ·
  Sales Channel · BU/Cost Centre · CM Level
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension


# ---------- styling ----------
FONT_NAME = "Arial"

COLOR_HEADER_BG = "1F3864"       # navy
COLOR_HEADER_FG = "FFFFFF"
COLOR_SUBHEAD_BG = "D9E1F2"      # light navy
COLOR_DOMAIN_BG = "2E75B6"       # med blue
COLOR_AXIS_BG = "E2EFDA"         # light green (axis block)
COLOR_FLAG_BG = "FFF59D"         # needs-lock yellow
COLOR_ZEBRA = "F2F2F2"
COLOR_BORDER = "BFBFBF"

thin = Side(style="thin", color=COLOR_BORDER)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_font(cell, bold=False, size=10, color="000000"):
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)


def style_header(cell, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FG):
    set_font(cell, bold=True, size=10, color=fg)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = ALIGN_HEADER
    cell.border = BORDER


def style_body(cell, bold=False, align=ALIGN_WRAP, fill=None):
    set_font(cell, bold=bold)
    cell.alignment = align
    cell.border = BORDER
    if fill:
        cell.fill = PatternFill("solid", start_color=fill)


# ---------- axes ----------
AXES = [
    ("Entity",       "INC / BNV / Holding / Consolidated"),
    ("Time",         "Weekly / Monthly / Quarterly / YTD / Annual"),
    ("Product",      "Macro Category (KSP/KRRS/KR) × Quality (Clear/Character) × Master/Profile"),
    ("Geography",    "Country (USA, BEL, DEU, NOR, ...)"),
    ("Sales Area",   "Global roll-up — US/CAN regions (North East, South East, Central, West, ...) + Europe regions (Scandinavia, DACH, Benelux, Southern EU, UK/IE, Rest of EU, ...) + Rest of World"),
    ("Sales Channel","Distributor / OEM / Industrial Application / Direct"),
    ("BU",           "10 Leadership · 15 Finance/IT · 20 Logistics · 21 Planning/CS · 30 Operations · 50 Sales Scand · 52 Sales Intl · 60 Marketing · 70 R&D"),
    ("CM Level",     "REV / COGS-D / COGS-I / SC / SALES / MKT / CORP / RD / ESOP / DA / FIN"),
]
# Note: Phase (Actual / BP / RF / LY) is captured in the "Phases (A/BP/RF/LY)" base column
# on every KPI row and in the BP — Readiness column; it is implicit on every KPI so not
# represented as a separate axis tick.
AXIS_KEYS = [a[0] for a in AXES]

# ---------- per-KPI column structure ----------
# Fixed columns then one tick column per axis
BASE_COLS = [
    ("#", 4),
    ("KPI Code", 12),
    ("KPI Name", 34),
    ("Definition", 55),
    ("Formula", 45),
    ("UoM", 10),
    ("Frequency", 12),
    ("Owner", 18),
    ("Op. Responsible", 20),
    ("Source System", 18),
    ("Actual — Readiness", 22),
    ("BP — Readiness", 18),
    ("Target / Threshold", 14),
    ("Phases (A/BP/RF/LY)", 15),
]
AXIS_COLS = [(a, 11) for a in AXIS_KEYS]
TAIL_COLS = [("Notes", 35)]

ALL_COLS = BASE_COLS + AXIS_COLS + TAIL_COLS


# ---------- KPI data ----------
# Each domain = list of rows. Columns must match ALL_COLS minus #.
# axes = list of axis keys that apply for THIS KPI (checkmarks placed on those).

def K(code, name, definition, formula, uom, freq, owner, opres, source, target, phases, axes, notes="", flag=False):
    return {
        "code": code, "name": name, "definition": definition, "formula": formula,
        "uom": uom, "freq": freq, "owner": owner, "opres": opres,
        "source": source, "target": target, "phases": phases,
        "axes": axes, "notes": notes, "flag": flag,
    }


FINANCE = [
    K("FIN-01", "EBITDA Margin", "Operating profitability before D&A and financial items, as % of revenue.",
      "EBITDA ÷ Revenue", "%", "Monthly", "CFO", "Group Controller", "Odoo GL",
      "Yes", "A / BP / RF / LY",
      ["Entity", "Time", "Geography", "BU", "CM Level"]),
    K("FIN-02", "EBITDA per m³", "Operating profit efficiency on a physical-unit basis. Cross-entity comparable.",
      "EBITDA ÷ m³ sold", "EUR/m³", "Monthly", "CFO", "Group Controller", "Odoo GL + Metrics",
      "Yes", "A / BP / RF / LY",
      ["Entity", "Time", "Product", "Geography", "Sales Channel"]),
    K("FIN-03", "GM I %", "Material margin: Revenue minus direct materials (wood FIFO + chemicals + packaging), as % of revenue.",
      "(Revenue − COGS-D) ÷ Revenue", "%", "Monthly", "CFO", "Group Controller", "Odoo analytics (CM Level)",
      "Yes", "A / BP / RF / LY",
      ["Entity", "Time", "Product", "Geography", "Sales Area", "Sales Channel", "CM Level"]),
    K("FIN-04", "GM II %", "Production margin: GM I minus indirect conversion (stacking, autoclave, dryer absorption, plant overhead).",
      "(Revenue − COGS-D − COGS-I) ÷ Revenue", "%", "Monthly", "CFO", "BNV Controller", "Odoo analytics (CM Level)",
      "Yes", "A / BP / RF / LY",
      ["Entity", "Time", "Product", "Geography", "CM Level"],
      notes="BNV only — INC is distribution so GM II = GM I."),
    K("FIN-05", "Revenue Growth YoY", "Revenue growth vs same period prior year, at constant FX.",
      "(Revenue − Revenue LY) ÷ Revenue LY", "%", "Monthly", "CFO", "Group Controller", "Odoo GL",
      "Yes", "A / LY",
      ["Entity", "Time", "Product", "Geography", "Sales Channel"]),
    K("FIN-06", "Payroll / Revenue", "Total personnel cost as % of revenue.",
      "Total Payroll ÷ Revenue", "%", "Monthly", "CFO", "HR + Finance", "Odoo payroll + GL",
      "Yes", "A / BP / RF",
      ["Entity", "Time", "BU"]),
    K("FIN-07", "Cost of Sales Payroll per m³", "Direct production payroll per unit of output produced.",
      "(Payroll BU=30 Operations) ÷ m³ produced", "EUR/m³", "Monthly", "CFO", "BNV Controller", "Odoo payroll + MES",
      "Yes", "A / BP / RF",
      ["Entity", "Time", "Product", "BU"]),
    K("FIN-08", "Operating Cash Flow", "Cash generated from operating activities, indirect method.",
      "EBITDA + ΔWorking Capital ± Non-cash items", "EUR", "Monthly", "CFO", "Group Controller", "Odoo CF statement",
      "Yes", "A / BP / RF",
      ["Entity", "Time"]),
    K("FIN-09", "Free Cash Flow (post-financing)", "Cash left after operating, investing AND debt service — the view that matters given Kebony's leverage. Financial cost is deliberately inside the FCF because interest + FX are material at Kebony.",
      "OCF − Capex − Net Financial Cost (interest paid + FX losses − interest received)", "EUR", "Monthly", "CFO", "Group Controller", "Odoo CF + Asset register + Treasury",
      "Yes", "A / BP / RF",
      ["Entity", "Time"],
      notes="Equivalent to FCF-to-Equity rather than FCF-to-Firm. Tracked alongside OCF and Capex so the financing drag is visible."),
    K("FIN-10", "DSO", "Average collection period for customer invoices.",
      "(AR ÷ Revenue) × period days", "days", "Monthly", "CFO", "Credit Manager", "Odoo AR ageing",
      "Yes", "A / BP / LY",
      ["Entity", "Time", "Geography", "Sales Channel", "Sales Area"]),
    K("FIN-11", "DPO", "Average payment period to suppliers.",
      "(AP ÷ Purchases) × period days", "days", "Monthly", "CFO", "AP Accountant", "Odoo AP ageing",
      "Yes", "A / BP",
      ["Entity", "Time"]),
    K("FIN-12", "DIO (Days Inventory Outstanding)", "Average days inventory is held before sale or consumption.",
      "(Inventory Value ÷ COGS) × period days", "days", "Monthly", "CFO", "Group Controller", "Odoo inventory + GL",
      "Yes", "A / BP",
      ["Entity", "Time", "Product"]),
    K("FIN-13", "Cash Conversion Cycle", "Total working-capital cycle in days — lower is better.",
      "DIO + DSO − DPO", "days", "Monthly", "CFO", "Group Controller", "Computed",
      "Yes", "A / BP",
      ["Entity", "Time"]),
    K("FIN-14", "Working Capital / Revenue", "Operating WC tied up per unit of revenue.",
      "(AR + Inventory − AP) ÷ Revenue", "%", "Monthly", "CFO", "Group Controller", "Odoo GL",
      "Yes", "A / BP",
      ["Entity", "Time"]),
    K("FIN-15", "Inventory Turns (annualised)", "How many times inventory rotates per year. Complement to DIO.",
      "Annual COGS ÷ Average Inventory (rolling 12M)", "x", "Monthly", "CFO", "Group Controller", "Odoo GL + inventory",
      "Yes", "A / BP / LY",
      ["Entity", "Time", "Product"]),
    K("FIN-16", "Net Debt / EBITDA", "Leverage — net financial debt over rolling-12 EBITDA.",
      "(Financial Debt − Cash) ÷ LTM EBITDA", "x", "Monthly", "CFO", "Group Controller", "Odoo GL",
      "Yes", "A / BP",
      ["Entity", "Time"]),
    K("FIN-17", "Bad Debt Ratio", "Provision for doubtful debt as % of AR.",
      "Bad Debt Provision ÷ AR", "%", "Monthly", "CFO", "Credit Manager", "Odoo AR",
      "Yes", "A / BP",
      ["Entity", "Time", "Geography"]),
]

SALES = [
    K("SAL-01", "Volume Sold (m³)", "Total volume delivered to end customers.",
      "Σ m³ delivered", "m³", "Weekly", "Head of Sales", "Sales Ops", "Odoo stock moves",
      "Yes", "A / BP / RF / LY",
      ["Entity", "Time", "Product", "Geography", "Sales Area", "Sales Channel"]),
    K("SAL-02", "Average Selling Price", "Net revenue per m³ — captures mix and price effect.",
      "Revenue ÷ m³ sold", "EUR/m³", "Monthly", "Head of Sales", "Sales Controller", "Odoo invoicing",
      "Yes", "A / BP / LY",
      ["Entity", "Time", "Product", "Geography", "Sales Channel"]),
    K("SAL-03", "OTD (On-Time Delivery)", "% of order lines delivered on or before committed date.",
      "Lines on-time ÷ Total lines shipped", "%", "Weekly", "Head of Sales", "Customer Service", "Odoo delivery",
      "Yes", "A",
      ["Entity", "Time", "Geography", "Sales Channel", "Sales Area"]),
    K("SAL-04", "OTIF (On-Time In-Full)", "% of orders delivered on time AND in full — stricter than OTD.",
      "Orders OTIF ÷ Total orders", "%", "Weekly", "Head of Sales", "Customer Service", "Odoo delivery",
      "Yes", "A",
      ["Entity", "Time", "Geography", "Sales Channel"]),
    K("SAL-05", "Order Backlog", "Open confirmed orders not yet delivered.",
      "Σ backlog (m³ or EUR)", "m³ / EUR", "Weekly", "Head of Sales", "Sales Ops", "Odoo SO",
      "Yes", "A / BP",
      ["Entity", "Time", "Product", "Geography", "Sales Channel"]),
    K("SAL-06", "Win Rate", "Commercial effectiveness — quotes that convert to orders.",
      "Orders Won ÷ Quotes Issued", "%", "Monthly", "Head of Sales", "Sales Ops", "Odoo CRM",
      "Yes", "A",
      ["Entity", "Time", "Geography", "Sales Channel"]),
    K("SAL-07", "Customer Concentration (Top 10)", "Revenue risk — share of top-10 customers in total revenue.",
      "Revenue Top-10 customers ÷ Total Revenue", "%", "Quarterly", "Head of Sales", "Sales Controller", "Odoo invoicing",
      "Yes (ceiling)", "A / LY",
      ["Entity", "Time", "Geography", "Sales Channel"]),
    K("SAL-08", "New Customer Acquisitions", "Count of customers first invoiced in the period.",
      "# partners first-time invoiced", "count", "Monthly", "Head of Sales", "Sales Ops", "Odoo invoicing",
      "Yes", "A / BP",
      ["Entity", "Time", "Geography", "Sales Channel"]),
    K("SAL-09", "New Customer Revenue Contribution", "Share of revenue from customers acquired in last 12 months.",
      "Revenue (new customers) ÷ Total Revenue", "%", "Monthly", "Head of Sales", "Sales Controller", "Odoo invoicing",
      "Yes", "A / LY",
      ["Entity", "Time", "Geography", "Sales Channel"]),
    K("SAL-10", "Churn Rate", "Customer attrition — buyers who bought last year but not this year.",
      "Lost customers (LY-A) ÷ LY customer count", "%", "Quarterly", "Head of Sales", "Sales Controller", "Odoo invoicing",
      "Yes", "A / LY",
      ["Entity", "Time", "Geography", "Sales Channel"]),
    K("SAL-11", "Consignment Volume Outbound", "US only — m³ shipped into consignment stock (not yet auto-invoiced).",
      "Σ m³ moved to consignment location", "m³", "Weekly", "Head of US", "US Controller", "Odoo stock moves",
      "Yes", "A / BP",
      ["Entity", "Time", "Product", "Sales Area", "Sales Channel"],
      notes="Ties to Outbound Consignment Requirements v0.8."),
    K("SAL-12", "Consignment Ageing > 60 days", "US only — share of consignment stock older than 60 days (near auto-invoice 90d).",
      "m³ consignment age > 60d ÷ Total consignment m³", "%", "Weekly", "Head of US", "US Controller", "Odoo stock ageing",
      "Yes (ceiling)", "A",
      ["Entity", "Time", "Product", "Sales Area"]),
    K("SAL-13", "Quote Cycle Time", "Speed of commercial response — request to quote sent.",
      "Avg(Quote date − Request date)", "days", "Monthly", "Head of Sales", "Sales Ops", "Odoo CRM",
      "Yes", "A",
      ["Entity", "Time", "Geography", "Sales Channel"]),
]

SUPPLY_CHAIN = [
    K("SC-01", "Supplier OTD (S-OTD)", "Share of supplier deliveries on or before committed date — inbound mirror of customer OTD.",
      "Lines on-time ÷ Total lines received", "%", "Monthly", "Head of SC", "Purchasing", "Odoo receipts",
      "Yes", "A",
      ["Entity", "Time", "Product"]),
    K("SC-02", "Supplier Lead Time (avg)", "Average lead time from PO confirmation to receipt.",
      "Avg(Receipt date − PO confirm date)", "days", "Monthly", "Head of SC", "Purchasing", "Odoo PO",
      "Yes", "A",
      ["Entity", "Time", "Product"]),
    K("SC-03", "Open PO Value", "Committed spend not yet received.",
      "Σ PO remaining value (open POs)", "EUR", "Weekly", "Head of SC", "Purchasing", "Odoo PO",
      "Yes", "A / BP",
      ["Entity", "Time", "Product"]),
    K("SC-04", "Freight Cost per m³", "Outbound freight cost per unit shipped — challenge 3PL and carriers.",
      "Freight cost ÷ m³ shipped", "EUR/m³", "Monthly", "Head of SC", "Logistics", "Odoo + carrier bills",
      "Yes", "A / BP",
      ["Entity", "Time", "Geography", "Sales Channel"]),
    K("SC-05", "Freight % of Revenue", "Outbound freight intensity.",
      "Freight cost ÷ Revenue", "%", "Monthly", "Head of SC", "Logistics", "Odoo GL",
      "Yes", "A / BP",
      ["Entity", "Time", "Geography"]),
    K("SC-06", "3PL / Warehousing Cost per m³", "Third-party storage + handling cost per unit stored.",
      "3PL cost ÷ avg m³ in 3PL location", "EUR/m³", "Monthly", "Head of SC", "Logistics", "Odoo GL + inventory",
      "Yes", "A / BP",
      ["Entity", "Time", "Geography"]),
    K("SC-07", "Stock-Out Events", "Number of SKU × location × week combinations with zero stock.",
      "# stock-out events", "count", "Weekly", "Head of SC", "Planning", "Odoo inventory",
      "Yes (ceiling)", "A",
      ["Entity", "Time", "Product"]),
    K("SC-08", "Forecast Accuracy (Demand)", "Planning quality — lower error = better replenishment.",
      "1 − Σ|Actual−Forecast| ÷ Σ|Actual|", "%", "Monthly", "Head of SC", "Planning", "Planning tool + Odoo",
      "Yes", "A",
      ["Entity", "Time", "Product", "Geography"]),
    K("SC-09", "Consignment Receivables (Biewer flow, US)", "Open consignment stock owed to Biewer before transfer-price invoicing.",
      "Σ Biewer consignment stock value", "USD", "Weekly", "Head of SC", "US Controller", "Odoo",
      "No — monitor", "A",
      ["Entity", "Time", "Product"]),
]

MANUFACTURING = [
    K("MFG-01", "OEE (Overall Equipment Effectiveness)", "Availability × Performance × Quality — composite WC efficiency.",
      "A × P × Q per work centre", "%", "Weekly", "Head of Mfg", "Plant Controller", "MES + Odoo MO",
      "Yes", "A / BP",
      ["Entity", "Time", "Product"],
      notes="Measured per work centre: Stacking, Autoclave, Dryer."),
    K("MFG-02", "Capacity Utilisation — Autoclave", "Scheduling efficiency of the autoclaves. Scheduled run hours vs theoretical available hours. EXCLUDES cycle-time and quality loss (those sit in OEE).",
      "Scheduled autoclave run hours ÷ (Shifts × Hrs/shift × Op. days − Planned maintenance)", "%", "Weekly", "Head of Mfg", "Plant Controller", "MES + Odoo shift calendar",
      "Yes", "A / BP",
      ["Entity", "Time"],
      notes="⚠ Definition must be locked with Head of Mfg — distinction from OEE is critical. Autoclave is the typical bottleneck on KSP capacity.", flag=True),
    K("MFG-03", "Capacity Utilisation — Dryer", "Scheduling efficiency of the dryers. Scheduled run hours vs theoretical available hours.",
      "Scheduled dryer run hours ÷ (Available DL slots × Hrs/DL − Planned maintenance)", "%", "Weekly", "Head of Mfg", "Plant Controller", "MES + Odoo DL calendar",
      "Yes", "A / BP",
      ["Entity", "Time"],
      notes="⚠ Definition must be locked with Head of Mfg. Dryer is the bottleneck for KRRS machined chain.", flag=True),
    K("MFG-04", "First Pass Yield (FPY)", "Share of output that passes quality gates first time without rework.",
      "Good output (first pass) ÷ Total output", "%", "Weekly", "Head of Mfg", "Quality + Plant", "MES + QC logs",
      "Yes", "A / BP",
      ["Entity", "Time", "Product"]),
    K("MFG-05", "Scrap Rate", "% of material scrapped during production.",
      "Scrap volume ÷ Input volume", "%", "Weekly", "Head of Mfg", "Plant Controller", "MES + Odoo MO",
      "Yes (ceiling)", "A / BP",
      ["Entity", "Time", "Product"]),
    K("MFG-06", "Dryer Cycle Time", "Average dryer load duration (schedule adherence).",
      "Avg(End date − Start date) per DL", "hours", "Weekly", "Head of Mfg", "Dryer Operator", "Odoo DL + MES",
      "Yes", "A",
      ["Entity", "Time", "Product"]),
    K("MFG-07", "Autoclave Cycle Time", "Average autoclave batch duration — by product (chemistry and species drive cycle length).",
      "Avg(End date − Start date) per batch", "hours", "Weekly", "Head of Mfg", "Autoclave Operator", "Odoo batch + MES",
      "Yes", "A",
      ["Entity", "Time", "Product"]),
    K("MFG-08", "Conversion Cost per m³", "Indirect cost absorbed per unit of output (COGS-I).",
      "Σ COGS-I ÷ m³ produced", "EUR/m³", "Monthly", "Head of Mfg", "Plant Controller", "Odoo analytics",
      "Yes", "A / BP",
      ["Entity", "Time", "Product", "BU"]),
    K("MFG-09", "Work Centre Absorption Variance", "Gap between actual and absorbed WC cost — triggers rate recalibration.",
      "Actual WC cost − Absorbed WC cost", "EUR", "Monthly", "Head of Mfg", "Plant Controller", "Odoo analytics",
      "Yes", "A / BP",
      ["Entity", "Time", "BU"]),
    K("MFG-10", "Unplanned Downtime", "Unplanned hours lost per work centre.",
      "Σ unplanned stops (hours)", "hours", "Weekly", "Head of Mfg", "Plant Operator", "MES",
      "Yes (ceiling)", "A",
      ["Entity", "Time"]),
]

QUALITY = [
    K("QUA-01", "Complaint Rate per m³", "Volume-normalised customer complaint frequency.",
      "Complaint count ÷ m³ sold × 1,000", "per 1,000m³", "Monthly", "Head of Quality", "Quality Mgr", "Claim log + Odoo",
      "Yes (ceiling)", "A",
      ["Entity", "Time", "Product", "Geography", "Sales Channel"]),
    K("QUA-02", "Cost of Poor Quality (CoPQ)", "Scrap + rework + returns + warranty as % of revenue.",
      "(Scrap + Rework + Returns + Warranty) ÷ Revenue", "%", "Monthly", "Head of Quality", "Quality + Finance", "Odoo GL",
      "Yes (ceiling)", "A / BP",
      ["Entity", "Time", "Product"]),
    K("QUA-03", "Reception Quality Gate Rejection", "% of inbound receipts rejected at Gate 1.",
      "# rejected receipts ÷ Total receipts", "%", "Monthly", "Head of Quality", "Reception", "Odoo receipts",
      "Yes (ceiling)", "A",
      ["Entity", "Time", "Product"]),
    K("QUA-04", "Production Quality Gate Rejection", "% of production runs rejected at Gate 2.",
      "# rejected runs ÷ Total runs", "%", "Monthly", "Head of Quality", "Quality Mgr", "MES + QC",
      "Yes (ceiling)", "A",
      ["Entity", "Time", "Product"]),
    K("QUA-05", "Pre-Dispatch Quality Gate Rejection", "% of packs rejected at Gate 3 (pre-ship).",
      "# rejected packs ÷ Total packs", "%", "Monthly", "Head of Quality", "Dispatch", "Odoo + QC",
      "Yes (ceiling)", "A",
      ["Entity", "Time", "Product"]),
    K("QUA-06", "2nd Choice Output %", "Share of finished goods routed to 2nd-choice location.",
      "m³ 2nd Choice ÷ m³ produced", "%", "Monthly", "Head of Quality", "Plant", "Odoo inventory",
      "Yes (ceiling)", "A / BP",
      ["Entity", "Time", "Product"]),
    K("QUA-07", "CAPA Closure Rate", "Corrective/preventive actions closed on time.",
      "CAPA closed on time ÷ CAPA due", "%", "Monthly", "Head of Quality", "Quality Mgr", "QC tool",
      "Yes", "A",
      ["Entity", "Time"]),
    K("QUA-08", "Customer Returns %", "Share of delivered volume returned by customers.",
      "m³ returned ÷ m³ sold", "%", "Monthly", "Head of Quality", "Customer Service", "Odoo RMA",
      "Yes (ceiling)", "A",
      ["Entity", "Time", "Product", "Geography", "Sales Channel"]),
    K("QUA-09", "Document Trail Completeness", "Share of shipments with full quality docs (cert, spec, batch traceability).",
      "Shipments with complete docs ÷ Total shipments", "%", "Monthly", "Head of Quality", "Dispatch + QA", "Odoo + docs",
      "Yes", "A",
      ["Entity", "Time", "Product", "Geography"],
      notes="Links to CEO ask: traceability reception → claim."),
]

INVENTORY = [
    K("INV-01", "DOS — Raw Wood", "Days of stock for raw wood (White Wood + Rough Sawn).",
      "m³ on hand ÷ avg daily consumption", "days", "Weekly", "Head of SC", "Planning", "Odoo inventory",
      "Yes", "A / BP",
      ["Entity", "Time", "Product"]),
    K("INV-02", "DOS — Chemicals", "Days of stock for Ready-Mix + components (FA, MA, CA, NaC).",
      "kg on hand ÷ avg daily consumption", "days", "Weekly", "Head of SC", "Planning", "Odoo inventory",
      "Yes", "A / BP",
      ["Entity", "Time"]),
    K("INV-03", "DOS — Finished Goods", "Days of stock for FG (KSP/KRRS/KR).",
      "m³ on hand ÷ avg daily sales", "days", "Weekly", "Head of SC", "Planning", "Odoo inventory",
      "Yes", "A / BP",
      ["Entity", "Time", "Product", "Geography"]),
    K("INV-04", "Dead / Slow-Moving Stock %", "FG with no movement for N months (N = 6) as % of total FG.",
      "m³ no movement ≥ 6M ÷ Total FG m³", "%", "Monthly", "Head of SC", "Planning", "Odoo inventory",
      "Yes (ceiling)", "A",
      ["Entity", "Time", "Product"]),
    K("INV-05", "Inventory Accuracy %", "Cycle count accuracy — discrepancy vs book value.",
      "1 − |Count − Book| ÷ Book", "%", "Monthly", "Head of SC", "Warehouse", "Cycle count logs",
      "Yes", "A",
      ["Entity", "Time"]),
]

SAFETY = [
    K("SAF-01", "LTI Count", "Number of Lost Time Injuries in period.",
      "# LTI events", "count", "Monthly", "HSE Manager", "Plant Safety", "HSE log",
      "Yes (ceiling)", "A / LY",
      ["Entity", "Time"]),
    K("SAF-02", "LTIR (Lost Time Injury Rate)", "LTI frequency per 1M hours worked — comparable across sites.",
      "LTI × 1,000,000 ÷ Hours worked", "per 1M h", "Monthly", "HSE Manager", "HR + HSE", "HSE log + HR",
      "Yes (ceiling)", "A / LY",
      ["Entity", "Time"]),
    K("SAF-03", "TRIR (Total Recordable Incident Rate)", "All recordable incidents (LTI + medical treatment) per 1M hours.",
      "Recordables × 1,000,000 ÷ Hours worked", "per 1M h", "Monthly", "HSE Manager", "HSE", "HSE log + HR",
      "Yes (ceiling)", "A / LY",
      ["Entity", "Time"]),
    K("SAF-04", "Severity Rate", "Days lost per 1,000 hours worked — measures severity not just frequency.",
      "Days lost × 1,000 ÷ Hours worked", "days/1k h", "Monthly", "HSE Manager", "HR + HSE", "HSE log + HR",
      "Yes (ceiling)", "A / LY",
      ["Entity", "Time"]),
    K("SAF-05", "Near Miss Count", "Reported near-miss events — leading indicator.",
      "# near miss reports", "count", "Monthly", "HSE Manager", "All employees", "HSE reporting tool",
      "Yes (floor — want more reports)", "A",
      ["Entity", "Time"],
      notes="Target is floor not ceiling: more reports = healthier safety culture."),
    K("SAF-06", "Days Without LTI", "Consecutive days since last LTI — leading / motivational indicator.",
      "Today − last LTI date", "days", "Monthly", "HSE Manager", "Plant Safety", "HSE log",
      "Yes (floor)", "A",
      ["Entity", "Time"]),
    K("SAF-07", "Safety Training Hours per FTE", "Average safety training delivered per head.",
      "Σ training hrs ÷ FTE", "hrs/FTE", "Quarterly", "HSE Manager", "HR + Training", "HR training records",
      "Yes", "A",
      ["Entity", "Time", "BU"]),
    K("SAF-08", "Safety Audits Completed", "Planned vs executed safety audits.",
      "Audits completed ÷ Audits planned", "%", "Quarterly", "HSE Manager", "HSE", "HSE tool",
      "Yes", "A",
      ["Entity", "Time"]),
]

ESG = [
    K("ESG-01", "Scope 1 Emissions", "Direct emissions from owned operations (fuel, gas, on-site combustion).",
      "Σ fuel × emission factor", "tCO2e", "Quarterly", "ESG Manager", "HSE + Finance", "Meter readings + factors",
      "Yes (ceiling)", "A / BP / LY",
      ["Entity", "Time"]),
    K("ESG-02", "Scope 2 Emissions", "Indirect emissions from purchased electricity / heat.",
      "kWh × grid emission factor", "tCO2e", "Quarterly", "ESG Manager", "Facilities", "Utility bills + grid factor",
      "Yes (ceiling)", "A / BP / LY",
      ["Entity", "Time"]),
    K("ESG-03", "Scope 3 Emissions", "Upstream/downstream emissions (materials, freight, end-of-life).",
      "Category × activity × factor (GHG Protocol)", "tCO2e", "Annual", "ESG Manager", "ESG + SC", "LCA + supplier data",
      "Yes (ceiling)", "A / LY",
      ["Entity", "Time", "Product"]),
    K("ESG-04", "Emissions Intensity", "Emissions per unit of output — decarbonisation trajectory.",
      "Scope 1+2 ÷ m³ produced", "kgCO2e/m³", "Quarterly", "ESG Manager", "ESG + Plant", "Computed",
      "Yes (ceiling)", "A / BP / LY",
      ["Entity", "Time", "Product"]),
    K("ESG-05", "Energy Intensity", "Energy used per unit produced.",
      "kWh consumed ÷ m³ produced", "kWh/m³", "Quarterly", "ESG Manager", "Facilities", "Utility bills + MES",
      "Yes (ceiling)", "A / BP",
      ["Entity", "Time"]),
    K("ESG-06", "Renewable Energy %", "Share of energy consumed from renewable sources.",
      "kWh renewable ÷ Total kWh", "%", "Quarterly", "ESG Manager", "Facilities", "Utility bills + PPAs",
      "Yes", "A / BP",
      ["Entity", "Time"]),
    K("ESG-07", "Water Intensity", "Water consumed per unit produced.",
      "m³ water ÷ m³ wood produced", "m³/m³", "Quarterly", "ESG Manager", "Facilities", "Meter readings + MES",
      "Yes (ceiling)", "A / BP",
      ["Entity", "Time"]),
    K("ESG-08", "Waste Intensity", "Waste generated per unit produced.",
      "kg waste ÷ m³ produced", "kg/m³", "Quarterly", "ESG Manager", "Plant + Facilities", "Waste manifests",
      "Yes (ceiling)", "A / BP",
      ["Entity", "Time", "Product"]),
    K("ESG-09", "Waste Diverted from Landfill", "Recycling + recovery rate.",
      "Waste recycled/recovered ÷ Total waste", "%", "Quarterly", "ESG Manager", "Facilities", "Waste manifests",
      "Yes", "A / BP",
      ["Entity", "Time"]),
    K("ESG-10", "Chemical Use Intensity", "Chemical kg consumed per m³ of FG produced.",
      "kg chemicals ÷ m³ produced", "kg/m³", "Quarterly", "ESG Manager", "Plant + Chemistry", "MES + Odoo",
      "Yes (ceiling)", "A / BP",
      ["Entity", "Time", "Product"]),
]

MARKETING = [
    K("MKT-01", "Marketing Spend / Revenue", "Marketing investment intensity.",
      "Marketing spend ÷ Revenue", "%", "Monthly", "Head of Marketing", "Marketing Controller", "Odoo + accruals",
      "Yes", "A / BP",
      ["Entity", "Time", "Geography", "Sales Channel"]),
    K("MKT-02", "Customer Acquisition Cost (CAC)", "Marketing + sales spend per new customer acquired.",
      "(Marketing + Sales spend) ÷ New customers", "EUR", "Quarterly", "Head of Marketing", "Mkt + Sales", "Odoo + CRM",
      "Yes (ceiling)", "A / BP",
      ["Entity", "Time", "Geography", "Sales Channel"]),
    K("MKT-03", "Marketing Qualified Leads (MQL)", "Leads meeting marketing qualification criteria.",
      "# MQLs", "count", "Monthly", "Head of Marketing", "Marketing Ops", "CRM",
      "Yes", "A / BP",
      ["Entity", "Time", "Geography", "Sales Channel"]),
    K("MKT-04", "Lead → Customer Conversion Rate", "Funnel effectiveness.",
      "New customers ÷ MQLs (lagged)", "%", "Quarterly", "Head of Marketing", "Mkt + Sales", "CRM",
      "Yes", "A",
      ["Entity", "Time", "Geography"]),
    K("MKT-05", "Campaign ROI", "Revenue uplift attributable to campaign vs campaign cost.",
      "(Attributed revenue − Cost) ÷ Cost", "x", "Per campaign", "Head of Marketing", "Marketing Controller", "CRM + Finance",
      "Yes", "A",
      ["Entity", "Geography", "Sales Channel"]),
    K("MKT-06", "Trade Show ROI", "Leads × conversion × avg order − event cost.",
      "(Attributed revenue − Event cost) ÷ Event cost", "x", "Per event", "Head of Marketing", "Event Mgr", "CRM + Finance",
      "Yes", "A",
      ["Entity", "Geography"]),
    K("MKT-07", "Brand Awareness Index", "Aided / unaided brand recall in target markets (survey).",
      "Survey score (0-100)", "index", "Annual", "Head of Marketing", "Marketing Ops", "Survey agency",
      "Yes", "A / LY",
      ["Entity", "Geography"]),
    K("MKT-08", "Website Traffic", "Unique monthly visitors to web properties.",
      "Unique visitors (GA/Matomo)", "count", "Monthly", "Head of Marketing", "Digital Mgr", "Analytics tool",
      "Yes", "A / LY",
      ["Entity", "Geography"]),
    K("MKT-09", "Distributor Fund Utilisation", "% of accrued distributor marketing fund spent vs balance.",
      "Spend ÷ (Accrual + Opening balance)", "%", "Monthly", "Head of Marketing", "Marketing Controller", "Odoo 215511",
      "Yes", "A",
      ["Entity", "Time", "Geography", "Sales Channel", "Sales Area"],
      notes="Ties to Kebony US accrual architecture (bucket on 215511)."),
]

PRODUCT_DEV = [
    K("PD-01", "New Products Launched", "Count of SKUs commercially launched in period.",
      "# new SKUs with first sale", "count", "Quarterly", "Head of R&D", "Product Manager", "Odoo + PLM",
      "Yes", "A / BP",
      ["Entity", "Time", "Product"]),
    K("PD-02", "NPD Revenue Contribution", "Share of revenue from products launched in last 3 years.",
      "Revenue (SKUs launched ≤ 3y) ÷ Total Revenue", "%", "Quarterly", "Head of R&D", "R&D + Sales Controller", "Odoo",
      "Yes", "A / LY",
      ["Entity", "Time", "Product", "Geography", "Sales Channel"]),
    K("PD-03", "Time to Market", "Avg duration concept → first commercial sale.",
      "Avg(First sale date − Concept date)", "months", "Quarterly", "Head of R&D", "R&D Project Mgr", "PLM + Odoo",
      "Yes (ceiling)", "A",
      ["Entity", "Product"]),
    K("PD-04", "R&D Spend / Revenue", "R&D intensity — investment in future products.",
      "R&D spend (BU=70) ÷ Revenue", "%", "Quarterly", "Head of R&D", "R&D Controller", "Odoo analytics",
      "Yes", "A / BP",
      ["Entity", "Time", "BU"]),
    K("PD-05", "R&D Pipeline Count", "Active concepts in progress (pre-launch).",
      "# open R&D projects", "count", "Monthly", "Head of R&D", "R&D Project Mgr", "PLM",
      "Yes", "A",
      ["Entity"]),
    K("PD-06", "SKU Rationalisation", "Net SKU change — trim dead SKUs, protect ROI of active.",
      "New SKUs − Retired SKUs", "count", "Quarterly", "Head of R&D", "Product Manager", "Odoo",
      "Yes", "A / BP",
      ["Entity", "Time", "Product"]),
]

CERTIFICATION = [
    K("CRT-01", "FSC Volume Coverage", "Volume certified under FSC as % of total volume sold.",
      "FSC-certified m³ ÷ Total m³ sold", "%", "Monthly", "Head of Quality", "QA + Supply Chain", "Odoo lot + certs",
      "Yes", "A",
      ["Entity", "Time", "Product", "Geography"],
      notes="Always paired with PEFC — see CRT-02 + dual coverage CRT-03."),
    K("CRT-02", "PEFC Volume Coverage", "Volume certified under PEFC as % of total volume sold.",
      "PEFC-certified m³ ÷ Total m³ sold", "%", "Monthly", "Head of Quality", "QA + Supply Chain", "Odoo lot + certs",
      "Yes", "A",
      ["Entity", "Time", "Product", "Geography"]),
    K("CRT-03", "FSC + PEFC Dual Coverage", "Volume covered by BOTH schemes — full-certification commercial claim.",
      "(FSC ∩ PEFC) m³ ÷ Total m³ sold", "%", "Monthly", "Head of Quality", "QA", "Odoo lot + certs",
      "Yes", "A",
      ["Entity", "Time", "Product", "Geography"]),
    K("CRT-04", "Certification Audits Passed", "Passed / completed audits (FSC, PEFC, Chain of Custody).",
      "Passed ÷ Completed audits", "%", "Annual", "Head of Quality", "QA", "Audit reports",
      "Yes (floor 100%)", "A",
      ["Entity"]),
    K("CRT-05", "Chain of Custody Traceability", "Share of sold volume with full CoC trace back to supplier lot.",
      "Traceable m³ ÷ Total m³ sold", "%", "Monthly", "Head of Quality", "QA + Odoo", "Odoo lot traceability",
      "Yes (floor)", "A",
      ["Entity", "Time", "Product"]),
]

HR = [
    K("HR-01", "Headcount (FTE)", "Full-time equivalent headcount at period close.",
      "Σ FTE", "FTE", "Monthly", "Head of HR", "HR Ops", "HR system",
      "Yes", "A / BP",
      ["Entity", "Time", "BU"]),
    K("HR-02", "Voluntary Turnover", "Annualised voluntary leavers — retention indicator.",
      "Voluntary leavers × 12 ÷ Avg headcount", "%", "Quarterly", "Head of HR", "HR Ops", "HR system",
      "Yes (ceiling)", "A / LY",
      ["Entity", "Time", "BU"]),
    K("HR-03", "Absenteeism Rate", "Unplanned absence days as % of planned working days.",
      "Absence days ÷ Planned working days", "%", "Monthly", "Head of HR", "HR Ops", "HR system",
      "Yes (ceiling)", "A",
      ["Entity", "Time", "BU"]),
    K("HR-04", "Training Hours per FTE", "Total training delivered per head (ex. safety).",
      "Σ training hrs ÷ FTE", "hrs/FTE", "Quarterly", "Head of HR", "HR + Training", "HR system",
      "Yes", "A",
      ["Entity", "Time", "BU"]),
    K("HR-05", "Time to Fill Open Positions", "Avg days from requisition open to offer accepted.",
      "Avg(Offer date − Requisition date)", "days", "Quarterly", "Head of HR", "Talent Acquisition", "HR system",
      "Yes (ceiling)", "A",
      ["Entity", "BU"]),
]


# Actual data readiness per KPI:
#   native  = already in standard Odoo / kebony_bol_report — no build
#   enrich  = exists in Odoo but needs custom field / compute / module extension
#   ext     = requires external system integration (MES, PLM, HSE tool, CRM, survey, facility meters, LCA)
#   manual  = captured by hand each period
ACTUAL_READINESS_MAP = {
    "native":  "✅ Native Odoo",
    "enrich":  "🟡 Odoo — enrichment",
    "ext":     "🔌 External system",
    "manual":  "✍️ Manual",
}
# BP readiness: ✅ Budgeted in cube form, ⊘ Monitor only (no BP value)
BP_READINESS_MAP = {
    "bp":      "✅ Budgeted",
    "target":  "🎯 Annual target",
    "mon":     "⊘ Monitor only",
}

# Per-KPI classification. Keyed by code.
READINESS = {
    # ---- Finance (all native — GL + analytics + metrics layer in place) ----
    "FIN-01": ("native",  "bp"),   # EBITDA margin
    "FIN-02": ("native",  "bp"),   # EBITDA per m³ — m³ from kebony_bol_report metrics
    "FIN-03": ("native",  "bp"),   # GM I % — CM Level analytic plan exists today
    "FIN-04": ("native",  "bp"),   # GM II % — CM Level exists; BNV BoM-split auto-post is P2 but computable
    "FIN-05": ("native",  "bp"),   # Revenue growth YoY
    "FIN-06": ("native",  "bp"),   # Payroll / Revenue
    "FIN-07": ("native",  "bp"),   # Cost of Sales Payroll per m³ — GL BU=30 + metrics
    "FIN-08": ("native",  "bp"),   # OCF — Odoo cash flow statement
    "FIN-09": ("native",  "bp"),   # FCF post-financing — interest in P&L, debt in BS
    "FIN-10": ("native",  "bp"),   # DSO
    "FIN-11": ("native",  "bp"),   # DPO
    "FIN-12": ("native",  "bp"),   # DIO
    "FIN-13": ("native",  "bp"),   # Cash Conversion Cycle
    "FIN-14": ("native",  "bp"),   # WC / Revenue
    "FIN-15": ("native",  "bp"),   # Inventory Turns
    "FIN-16": ("native",  "bp"),   # Net Debt / EBITDA — debt accounts + GL
    "FIN-17": ("native",  "bp"),   # Bad Debt Ratio

    # ---- Sales (all native once quotation discipline enforced + customer data in place) ----
    "SAL-01": ("native",  "bp"),   # Volume m³ — metrics layer
    "SAL-02": ("native",  "bp"),   # ASP
    "SAL-03": ("native",  "target"),  # OTD — delivery commitment vs done date (Odoo native)
    "SAL-04": ("native",  "target"),  # OTIF
    "SAL-05": ("native",  "bp"),   # Backlog
    "SAL-06": ("native",  "target"),  # Win Rate — quotation → order flow if discipline enforced
    "SAL-07": ("native",  "mon"),  # Customer concentration
    "SAL-08": ("native",  "bp"),   # New customer acquisitions — first-invoice logic on partner
    "SAL-09": ("native",  "bp"),   # New customer revenue % — same base
    "SAL-10": ("native",  "mon"),  # Churn rate — LY vs A customer base
    "SAL-11": ("native",  "bp"),   # Consignment volume — location-based stock moves
    "SAL-12": ("native",  "target"),  # Consignment ageing — kebony_bol_report pack reservation has ageing
    "SAL-13": ("native",  "target"),  # Quote cycle time — Odoo CRM/Sales dates

    # ---- Supply Chain ----
    "SC-01":  ("native",  "target"),  # S-OTD — PO confirm date vs receipt
    "SC-02":  ("native",  "mon"),     # Supplier lead time
    "SC-03":  ("native",  "bp"),      # Open PO value
    "SC-04":  ("native",  "bp"),      # Freight per m³ — GL + metrics
    "SC-05":  ("native",  "bp"),      # Freight % revenue
    "SC-06":  ("native",  "bp"),      # 3PL / warehousing per m³
    "SC-07":  ("enrich",  "target"),  # Stock-out events — needs event-capture logic beyond standard report
    "SC-08":  ("ext",     "target"),  # Forecast accuracy — planning tool outside ERP
    "SC-09":  ("native",  "mon"),     # Biewer consignment — kebony_bol_report handles this today

    # ---- Manufacturing (MES gaps only where precision needed) ----
    "MFG-01": ("ext",     "target"),  # OEE — MES (A × P × Q precision)
    "MFG-02": ("ext",     "target"),  # Capacity Util Autoclave — MES run hours
    "MFG-03": ("ext",     "target"),  # Capacity Util Dryer — MES + Odoo DL calendar
    "MFG-04": ("ext",     "target"),  # FPY — MES + QC
    "MFG-05": ("native",  "bp"),      # Scrap rate — Odoo MO tracks scrap natively
    "MFG-06": ("native",  "mon"),     # Dryer cycle — kebony_manufacturing DL start/end
    "MFG-07": ("native",  "mon"),     # Autoclave cycle — batch start/end on model
    "MFG-08": ("native",  "bp"),      # Conversion cost / m³ — CM Level + metrics
    "MFG-09": ("native",  "bp"),      # WC absorption variance — analytics native
    "MFG-10": ("ext",     "target"),  # Unplanned downtime — MES

    # ---- Quality ----
    "QUA-01": ("ext",     "target"),  # Complaint rate — claim log module needed; m³ is native
    "QUA-02": ("native",  "bp"),      # CoPQ — GL + analytics decomposition
    "QUA-03": ("enrich",  "target"),  # Reception QG — quarantine flag / quality alert integration
    "QUA-04": ("ext",     "target"),  # Production QG — MES/QC
    "QUA-05": ("enrich",  "target"),  # Pre-dispatch QG — quality check module integration
    "QUA-06": ("native",  "target"),  # 2nd choice % — location-based moves (Stock Locations v2.0 has 2nd Choice location)
    "QUA-07": ("ext",     "target"),  # CAPA — QC tool
    "QUA-08": ("native",  "target"),  # Customer returns % — Odoo RMA + metrics
    "QUA-09": ("ext",     "target"),  # Document trail — DMS

    # ---- Inventory (all native — stock valuation + metrics layer) ----
    "INV-01": ("native",  "bp"),      # DOS raw wood
    "INV-02": ("native",  "bp"),      # DOS chemicals
    "INV-03": ("native",  "bp"),      # DOS FG
    "INV-04": ("native",  "target"),  # Dead / slow-moving % — no-movement query
    "INV-05": ("enrich",  "target"),  # Inventory accuracy — cycle count discipline needed

    # ---- Safety (all external — HSE log not in Odoo) ----
    "SAF-01": ("ext",     "target"),  # LTI count
    "SAF-02": ("ext",     "target"),  # LTIR
    "SAF-03": ("ext",     "target"),  # TRIR
    "SAF-04": ("ext",     "target"),  # Severity
    "SAF-05": ("ext",     "mon"),     # Near miss
    "SAF-06": ("ext",     "mon"),     # Days without LTI
    "SAF-07": ("native",  "target"),  # Safety training hrs / FTE — Odoo HR training module (to be implemented)
    "SAF-08": ("ext",     "target"),  # Safety audits

    # ---- ESG (external — meters, LCA, utility bills) except chemical (internal BoM) ----
    "ESG-01": ("ext",     "target"),  # Scope 1
    "ESG-02": ("ext",     "target"),  # Scope 2
    "ESG-03": ("ext",     "target"),  # Scope 3 — LCA
    "ESG-04": ("ext",     "target"),  # Emissions intensity (depends on 1+2)
    "ESG-05": ("ext",     "target"),  # Energy intensity
    "ESG-06": ("ext",     "target"),  # Renewable %
    "ESG-07": ("ext",     "target"),  # Water intensity
    "ESG-08": ("ext",     "target"),  # Waste intensity
    "ESG-09": ("ext",     "target"),  # Waste diverted
    "ESG-10": ("native",  "target"),  # Chemical use intensity — BoM consumption + MO + metrics

    # ---- Marketing ----
    "MKT-01": ("native",  "bp"),      # Marketing spend / rev
    "MKT-02": ("native",  "target"),  # CAC — new customer count + marketing spend
    "MKT-03": ("native",  "target"),  # MQL — Odoo CRM leads (to be implemented)
    "MKT-04": ("native",  "target"),  # Lead → Customer conversion — Odoo CRM opportunity→partner
    "MKT-05": ("native",  "target"),  # Campaign ROI — Odoo Marketing campaigns (to be implemented)
    "MKT-06": ("native",  "target"),  # Trade show ROI — tracked as campaign in Odoo Marketing
    "MKT-07": ("manual",  "target"),  # Brand awareness — survey agency
    "MKT-08": ("ext",     "target"),  # Website traffic — GA/Matomo (separate tool)
    "MKT-09": ("native",  "bp"),      # Distributor fund utilisation — account 215511 exists

    # ---- Product Dev ----
    "PD-01":  ("native",  "bp"),      # New products launched — first-sale date derivable
    "PD-02":  ("native",  "target"),  # NPD revenue %
    "PD-03":  ("ext",     "target"),  # Time to market — PLM
    "PD-04":  ("native",  "bp"),      # R&D / rev (BU=70)
    "PD-05":  ("ext",     "mon"),     # R&D pipeline — PLM
    "PD-06":  ("native",  "target"),  # SKU rationalisation — product list

    # ---- Certification (lot-level cert blueprint committed) ----
    "CRT-01": ("native",  "target"),  # FSC volume coverage — lot-level cert tracking planned
    "CRT-02": ("native",  "target"),  # PEFC volume coverage
    "CRT-03": ("native",  "target"),  # Dual coverage
    "CRT-04": ("manual",  "target"),  # Audits passed — audit agency reports
    "CRT-05": ("native",  "target"),  # CoC traceability — Odoo lot traceability native

    # ---- HR (Odoo HR to be implemented — headcount, timekeeping, training) ----
    "HR-01":  ("native",  "bp"),      # FTE — Odoo HR employees
    "HR-02":  ("native",  "target"),  # Voluntary turnover — contract end date + leaver reason
    "HR-03":  ("native",  "target"),  # Absenteeism — time off + attendance
    "HR-04":  ("native",  "target"),  # Training hours / FTE — Odoo training module
    "HR-05":  ("ext",     "target"),  # Time to fill — requires Odoo Recruitment (not in scope)
}


DOMAINS = [
    ("Finance",          "CFO",                 FINANCE,       "1F3864"),
    ("Sales",            "Head of Sales / Head of US", SALES,  "2E75B6"),
    ("Supply Chain",     "Head of SC",          SUPPLY_CHAIN,  "548235"),
    ("Manufacturing",    "Head of Mfg",         MANUFACTURING, "8B4513"),
    ("Quality",          "Head of Quality",     QUALITY,       "7030A0"),
    ("Inventory",        "Head of SC",          INVENTORY,     "BF8F00"),
    ("Safety",           "HSE Manager",         SAFETY,        "C00000"),
    ("ESG",              "ESG Manager",         ESG,           "1E6B52"),
    ("Marketing",        "Head of Marketing",   MARKETING,     "E97132"),
    ("Product Dev",      "Head of R&D",         PRODUCT_DEV,   "636363"),
    ("Certification",    "Head of Quality",     CERTIFICATION, "005A9C"),
    ("HR",               "Head of HR",          HR,            "833C0C"),
]


# ---------- builder ----------
def build_domain_sheet(wb, name, owner_title, kpis, domain_color):
    ws = wb.create_sheet(name)

    # Title band
    total_cols = len(ALL_COLS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title = ws.cell(row=1, column=1, value=f"KEBONY KPI GOVERNANCE  —  {name.upper()}")
    set_font(title, bold=True, size=14, color="FFFFFF")
    title.fill = PatternFill("solid", start_color=domain_color)
    title.alignment = ALIGN_CENTER

    # Subtitle band (owner + signoff)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    sub = ws.cell(row=2, column=1,
                  value=f"Domain Owner: {owner_title}     |     "
                        f"Signoff (name + date): _______________________________     |     "
                        f"Please review each KPI — definition, formula, owner, axes.")
    set_font(sub, bold=True, size=10)
    sub.fill = PatternFill("solid", start_color=COLOR_SUBHEAD_BG)
    sub.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Header row 1 — group bands
    # Cols 1..12 = base, 13..13+len(AXES)-1 = axes, then tail
    # Row 3: group header; Row 4: per-column header
    # Base group
    base_start, base_end = 1, len(BASE_COLS)
    axis_start = base_end + 1
    axis_end = axis_start + len(AXIS_COLS) - 1
    tail_start = axis_end + 1
    tail_end = tail_start + len(TAIL_COLS) - 1

    ws.merge_cells(start_row=3, start_column=base_start, end_row=3, end_column=base_end)
    c = ws.cell(row=3, column=base_start, value="IDENTIFICATION · GOVERNANCE · METHOD")
    style_header(c, bg=COLOR_HEADER_BG)

    ws.merge_cells(start_row=3, start_column=axis_start, end_row=3, end_column=axis_end)
    c = ws.cell(row=3, column=axis_start, value="AXES OF ANALYSIS  (tick if this KPI slices on this dimension)")
    style_header(c, bg="548235")  # green band for cube axes

    ws.merge_cells(start_row=3, start_column=tail_start, end_row=3, end_column=tail_end)
    c = ws.cell(row=3, column=tail_start, value="COMMENTS")
    style_header(c, bg=COLOR_HEADER_BG)

    # Row 4: column headers
    col = 1
    for hdr, w in ALL_COLS:
        cell = ws.cell(row=4, column=col, value=hdr)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = w
        col += 1

    # Body rows start at 5
    row = 5
    # Readiness-cell palette
    READINESS_FILLS = {
        "✅ Native Odoo":          "C6EFCE",
        "🟡 Odoo — enrichment":    "FFEB9C",
        "🔌 External system":      "FCE4D6",
        "✍️ Manual":               "E7E6E6",
        "✅ Budgeted":             "C6EFCE",
        "🎯 Annual target":        "FFEB9C",
        "⊘ Monitor only":          "E7E6E6",
    }

    for i, k in enumerate(kpis, start=1):
        fill = COLOR_FLAG_BG if k["flag"] else (COLOR_ZEBRA if i % 2 == 0 else None)

        # Readiness lookup
        act_tag, bp_tag = READINESS.get(k["code"], ("manual", "mon"))
        actual_str = ACTUAL_READINESS_MAP[act_tag]
        bp_str = BP_READINESS_MAP[bp_tag]

        # Base columns (now with two readiness cells inserted before Target)
        vals = [
            i, k["code"], k["name"], k["definition"], k["formula"], k["uom"],
            k["freq"], k["owner"], k["opres"], k["source"],
            actual_str, bp_str,
            k["target"], k["phases"],
        ]
        # Center-aligned column indices (0-based): #, UoM, Freq, Actual, BP, Target, Phases
        CENTER_COLS = {0, 5, 6, 10, 11, 12, 13}
        for j, v in enumerate(vals):
            c = ws.cell(row=row, column=j + 1, value=v)
            extra_fill = None
            if j == 10:
                extra_fill = READINESS_FILLS.get(actual_str)
            elif j == 11:
                extra_fill = READINESS_FILLS.get(bp_str)
            style_body(c,
                       align=(ALIGN_CENTER if j in CENTER_COLS else ALIGN_WRAP),
                       fill=extra_fill if extra_fill else fill)
            if j in (10, 11):
                set_font(c, bold=True, size=9)

        # Axis checkboxes
        for j, axis in enumerate(AXIS_KEYS):
            v = "✓" if axis in k["axes"] else ""
            c = ws.cell(row=row, column=axis_start + j, value=v)
            style_body(c, align=ALIGN_CENTER, fill=COLOR_AXIS_BG if v else fill)
            if v:
                set_font(c, bold=True, color="1E6B52")

        # Notes
        c = ws.cell(row=row, column=tail_start, value=k["notes"])
        style_body(c, align=ALIGN_WRAP, fill=fill)

        # Row height hint for wrap
        ws.row_dimensions[row].height = 60
        row += 1

    # Freeze first 4 rows + first 3 cols (#, code, name)
    ws.freeze_panes = "D5"

    # Page setup for print
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.print_options.horizontalCentered = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def build_cover(wb):
    ws = wb.create_sheet("Cover", 0)

    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="KEBONY — KPI GOVERNANCE TEMPLATE")
    set_font(t, bold=True, size=18, color="FFFFFF")
    t.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    t.alignment = ALIGN_CENTER

    ws.merge_cells("A2:H2")
    s = ws.cell(row=2, column=1, value="v2.0  ·  21 April 2026  ·  VVF Consulting  ·  For domain-head review and signoff")
    set_font(s, size=11)
    s.alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 40
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 20

    # How to use
    ws.cell(row=4, column=1, value="HOW TO USE THIS TEMPLATE")
    set_font(ws.cell(row=4, column=1), bold=True, size=13, color=COLOR_HEADER_BG)
    ws.merge_cells("A4:H4")

    instructions = [
        "1.  Each domain has its own sheet. Open the sheet for your domain.",
        "2.  Review every KPI row: definition, formula, owner, operational responsible, frequency.",
        "3.  Check the AXES OF ANALYSIS columns (green band) — these are the dimensions of the management cube along which each KPI will be sliced. See the 'Axes of Analysis' reference sheet for the full list of values per axis.",
        "4.  Add / remove / reword KPIs directly in the sheet. Use the COMMENTS column for rationale.",
        "5.  Sign off in the yellow block below (or in the sheet header) and return the file.",
        "6.  Only DERIVED KPIs are listed (ratios, indices, computed values). Raw inputs (revenue, m³, payroll …) are implicit in the formula.",
        "7.  Target/Threshold column: 'Yes' if we set a target; 'No — analytical' if it is purely for analysis; include (ceiling) or (floor) when the direction matters.",
        "8.  Frequency column: reporting cadence (Daily / Weekly / Monthly / Quarterly / Annual / Per event).",
        "9.  Phases: which versions of the cube this KPI applies to — A (Actual), BP (Budget Plan), RF (Reforecast), LY (Last Year).",
        "10. The template will become an Odoo module (Cube Architecture — see Obsidian vault).",
    ]
    r = 5
    for line in instructions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(row=r, column=1, value=line)
        set_font(c, size=11)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="DOMAIN SIGN-OFF TRACKER")
    set_font(ws.cell(row=r, column=1), bold=True, size=13, color=COLOR_HEADER_BG)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    # sign-off table
    hdrs = ["Domain", "Owner Title", "Signoff Name", "Date", "Status", "# KPIs", "", ""]
    for j, h in enumerate(hdrs):
        c = ws.cell(row=r, column=j + 1, value=h)
        if h:
            style_header(c)
    r += 1
    for name, owner, kpis, _color in DOMAINS:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=owner)
        ws.cell(row=r, column=3, value="")
        ws.cell(row=r, column=4, value="")
        ws.cell(row=r, column=5, value="⬜ Pending")
        ws.cell(row=r, column=6, value=len(kpis))
        for col in range(1, 7):
            style_body(ws.cell(row=r, column=col),
                       align=ALIGN_CENTER if col in (4, 5, 6) else ALIGN_WRAP,
                       fill=COLOR_FLAG_BG if col == 5 else None)
        r += 1

    # Total row
    total = sum(len(d[2]) for d in DOMAINS)
    ws.cell(row=r, column=1, value="TOTAL")
    set_font(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=6, value=total)
    set_font(ws.cell(row=r, column=6), bold=True)
    for col in range(1, 7):
        style_body(ws.cell(row=r, column=col), align=ALIGN_CENTER)


def build_axes_sheet(wb):
    ws = wb.create_sheet("Axes of Analysis", 1)

    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value="KEBONY REPORTING CUBE — AXES OF ANALYSIS")
    set_font(t, bold=True, size=16, color="FFFFFF")
    t.fill = PatternFill("solid", start_color="548235")
    t.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:E2")
    c = ws.cell(row=2, column=1,
                value="Every KPI slices on one or more of the axes below. The cube is the SAME structure for Actuals (from Odoo SQL views) and for Budget (from input forms). See vault docs: 'Reporting - Cube Architecture' and 'Analytical Accounting Architecture'.")
    set_font(c, size=10)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[2].height = 45

    # Column widths
    widths = [22, 18, 42, 48, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Headers
    hdrs = ["Axis", "Cube dim code", "Values", "Business meaning / assignment rule", "Mandatory?"]
    for j, h in enumerate(hdrs):
        c = ws.cell(row=4, column=j + 1, value=h)
        style_header(c, bg="548235")

    axis_rows = [
        ("Entity",         "entity",     "INC · BNV · Holding · Consolidated",
         "Legal company. INC = Kebony Inc. (US, USD), BNV = production (BE, EUR), Holding = corporate (BE, EUR). CONS = intercompany-eliminated view.",
         "Yes (all KPIs)"),
        ("Time",           "period",     "Weekly (leaf) · Monthly (leaf) · Quarterly · YTD · Annual",
         "Two leaf-level grains kept in parallel: Monthly (financial KPIs, rolls up to Q / YTD / Year) and Weekly (operational KPIs — OTD, stock, production). Weeks do not nest into months (ISO weeks straddle month boundaries), so they are a separate storage level, not a sub-level of monthly.",
         "Yes (all KPIs)"),
        ("Phase / Version","version",    "Actual · BP (Budget Plan) · RF1 · RF2 · Last Year",
         "Actuals auto-computed from Odoo SQL. Budget entered via input forms. Variance never stored — computed (A vs BP, A vs LY).",
         "Yes (all KPIs)"),
        ("Product",        "product",    "Macro Category (KSP = Scots Pine · KRRS = Radiata Rough Sawn · KR = Radiata Machined) → Quality (Clear / Character) → Master / Profile",
         "Macro Category is a processing-state × species grouping, not pure species (KRRS and KR are both Radiata, at different processing stages). Derived from `x_studio_product_group` for the macro level and Product Master for drill-down.",
         "When applicable"),
        ("Geography",      "country",    "USA · CAN · BEL · DEU · NOR · FRA · GBR · 70+ others",
         "Country of end customer (shipping address on invoice). For IC sales BNV→INC, tagged USA (Canada gap fine-tuned at EOM).",
         "Yes for P&L lines"),
        ("Sales Area",     "sales_area", "US/CAN:  North East · South East · Central · Midwest · West · N. Rockies · OEM · East Canada · West Canada     |     Europe:  Scandinavia · DACH · Benelux · Southern EU · UK/IE · Eastern EU · Rest of EU     |     Rest of World (values to lock with Heads of Sales)",
         "Global roll-up — commercial segmentation, NOT geography. One Country can split across Sales Areas (e.g. USA → North East + South East + Central ...); one Sales Area can span Countries (e.g. DACH = DEU + AUT + CHE; Benelux = NLD + BEL + LUX). Set on `res.partner.x_kebony_sales_area`, inherits to SO → Invoice → COGS.",
         "Yes (all entities)"),
        ("Sales Channel",  "channel",    "Distributor · OEM · Industrial Application · Direct",
         "Commercial route-to-market. Distributor = resells our brand; OEM = integrates our wood into their branded product; Industrial Application = bulk/contract use (e.g. cladding for infrastructure, marine, large-scale façade projects); Direct = end-user sales. Drives marketing accrual routing (central vs distributor fund) and margin profile per channel.",
         "When applicable"),
        ("BU / Cost Centre","bu",        "10 Leadership · 15 Finance/IT · 20 Logistics · 21 Planning/CS · 30 Operations · 50 Sales Scandinavia · 52 Sales Intl · 60 Marketing · 70 R&D",
         "Hybrid business unit + cost centre. Drives payroll allocation, department P&L, CM-level default mapping.",
         "Yes for P&L lines"),
        ("CM Level",       "cm_level",   "REV · COGS-D · COGS-I · SC · SALES · MKT · CORP · RD · ESOP · DA · FIN",
         "Position in the P&L waterfall. Every GL account maps to exactly one CM Level — drives GM I, GM II, EBITDA decomposition.",
         "Yes for P&L lines"),
    ]
    r = 5
    for row in axis_rows:
        for j, v in enumerate(row):
            c = ws.cell(row=r, column=j + 1, value=v)
            style_body(c, align=(ALIGN_WRAP if j in (2, 3) else ALIGN_CENTER))
        ws.row_dimensions[r].height = 62
        r += 1

    # Matrix block
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value="HOW THE CUBE WORKS — STAR SCHEMA")
    set_font(c, bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="548235")
    c.alignment = ALIGN_CENTER
    r += 1

    block = [
        "FACT TABLE  =  one row per  (Account × Entity × Period × Version × Dimension × UoM)  →  VALUE",
        "ACCOUNT is a UNIFIED chart-of-accounts: financial accounts (REV/COGS/EBITDA), volume accounts (m³, LF, boards), operational KPIs (OTIF, DSO, scrap), cash flow, people.",
        "UoM is a first-class dimension: same KPI can be in EUR, USD, m³, LF, %, days, FTE simultaneously.",
        "Two data feeds: (a) Actuals — nightly SQL views from Odoo; (b) Budget — manual entry through Odoo forms. They never mix.",
    ]
    for line in block:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(row=r, column=1, value=line)
        set_font(c, size=10)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1

    ws.freeze_panes = "A5"


def build_legend(wb):
    ws = wb.create_sheet("Legend")

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="LEGEND & CONVENTIONS")
    set_font(t, bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    t.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 30

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 80

    entries = [
        ("Item",                 "Symbol / colour",  "Meaning"),
        ("Axis checkmark",       "✓ (green)",        "KPI slices on this dimension — cube query will use it as a pivot axis."),
        ("Yellow-highlighted row", "yellow fill",   "Definition or formula requires validation / lock by domain head."),
        ("Target direction: (ceiling)", "Target column", "Lower is better (e.g. scrap, complaints, DSO, churn)."),
        ("Target direction: (floor)",   "Target column", "Higher is better (e.g. OTD, FPY, safety days without LTI, near-miss reports)."),
        ("Frequency",            "—",                 "Daily / Weekly / Monthly / Quarterly / Annual / Per event / Per campaign."),
        ("Phase code 'A'",       "—",                 "Actual."),
        ("Phase code 'BP'",      "—",                 "Budget Plan (annual)."),
        ("Phase code 'RF'",      "—",                 "Reforecast (mid-year + Q3 revision)."),
        ("Phase code 'LY'",      "—",                 "Last Year Actual (for YoY variance)."),
        ("Derived KPI only",     "—",                 "Elementary/raw data (revenue, volumes, payroll …) is shown in the FORMULA column, not as separate KPI rows."),
        ("Domain owner",         "—",                 "Accountable for the KPI reporting discipline. Signs off on the definition."),
        ("Op. responsible",      "—",                 "Produces / publishes the KPI monthly. Usually a controller or ops manager."),
        ("Source system",        "—",                 "Where the data comes from: Odoo GL, Odoo analytics, MES, HR system, CRM, external survey, etc."),
    ]
    r = 3
    for row in entries:
        bold = (r == 3)
        for j, v in enumerate(row):
            c = ws.cell(row=r, column=j + 1, value=v)
            if bold:
                style_header(c)
            else:
                style_body(c, align=ALIGN_WRAP)
        ws.row_dimensions[r].height = 30 if not bold else 22
        r += 1


def build():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_cover(wb)
    build_axes_sheet(wb)

    for name, owner, kpis, color in DOMAINS:
        build_domain_sheet(wb, name, owner, kpis, color)

    build_legend(wb)

    # Set default font on workbook defaults (cosmetic)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    out = "/Users/oeb/Documents/Manufacturing Kebony/06 - Slide Decks/Kebony_KPI_Governance_Template.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"Domains: {len(DOMAINS)}")
    print(f"Total KPIs: {sum(len(d[2]) for d in DOMAINS)}")
    for name, _o, kpis, _c in DOMAINS:
        print(f"  {name}: {len(kpis)} KPIs")


if __name__ == "__main__":
    build()
